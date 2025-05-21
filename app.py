import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ====== App UI ======
st.title("🧭 靈數日曆 Excel 優化工具")
st.markdown("將上傳的日曆檔案中，\n- ✅ 所有儲存格垂直與水平置中\n- ✅ 將英文星期轉為中文格式\n- ✅ 表頭加粗並上底色\n- ✅ 每列加高行距")

uploaded_file = st.file_uploader("📤 請上傳日曆 Excel 檔案（.xlsx）", type=["xlsx"])

if uploaded_file:
    st.success("✅ 檔案上傳成功！開始進行優化處理…")

    # ====== Excel 處理邏輯 ======
    wb = load_workbook(uploaded_file)
    ws = wb.active

    # 替換英文星期為中文
    weekday_map = {
        "Sunday": "週日", "Monday": "週一", "Tuesday": "週二",
        "Wednesday": "週三", "Thursday": "週四",
        "Friday": "週五", "Saturday": "週六"
    }

    # 取得欄位名稱，找出「星期」欄位 index
    header = [cell.value for cell in ws[1]]
    week_col_idx = header.index("星期") + 1 if "星期" in header else None

    # 標題列樣式（加粗、底色）
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)

    # 處理每一列資料（含標題列）
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row)):
        ws.row_dimensions[i + 1].height = 22  # 行高設定
        for j, cell in enumerate(row):
            # 所有儲存格置中（含自動換行）
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # 標題列樣式
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font

        # 替換「星期」欄位為中文
        if i >= 1 and week_col_idx:
            cell = row[week_col_idx - 1]
            if cell.value in weekday_map:
                cell.value = weekday_map[cell.value]

    # 儲存為 BytesIO 並提供下載
    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    st.download_button(
        label="📥 下載優化後的 Excel",
        data=output,
        file_name="LuckyCalendar_優化版.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
